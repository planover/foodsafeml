# -*- coding: utf-8 -*-
"""
FoodSafeML (foodsafeml) v1.11.0 — 工作簿生成器（全量矩阵 + 权威数据填充版）
==========================================================
给定某食品目录下的 basic.json / translations.json / risks.json，生成
<食品名>_食品安全风险识别表.xlsx（3 个工作表）。

v1.11.0 核心升级（国际基准回填 + 指标压实）：
  ① 地区清单 235 个（全球主权国家+地区+国际组织/联盟/论坛，按洲分组）。
  ② 内置 BASE_LIMITS 权威限量数据库扩充至 130+ 条：Codex / EU / US / JP / CN
     五套基准真实联网检索结果，有数据的填实际值+来源URL，大幅减少 [待填写]。
  ③ 标准体系映射回填：采纳/参照 Codex/EU/US/JP 的国家与地区，直接回填对应
     基准的具体限量值（绿色数据 + 外链法规原文），而非仅标注"适用标准框架"。
  ④ 每条数据带 source_url 来源网址，可追溯法规原文。
  ⑤ Sheet4 全量矩阵：126指标 × 235地区，按洲分区，冻结首行首列。

防幻觉规则：缺失的多语字段/具体物质一律以 [待填写] 标注，绝不编造；
但优先用 BASE_LIMITS 权威数据 + 标准体系映射填充，最大化减少空白。

用法：
    python build_xlsx.py <食品目录> [<食品中文名>]
例：
    python build_xlsx.py ./food-safety-芒果干 芒果干
"""
import os
import re
import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SKILL_VER = "1.11.0"

# ====== 全量地区数据（235 个：国家/地区 + 国际组织）来自 regions_data 模块 ======
from regions_data import (
    REGION_INFO, REGION_FULLNAME, REGION_SYSTEM, REGION_CONTINENT, REGION_NOTE,
    ALL_REGIONS, REGIONS_BY_CONTINENT, CONTINENT_ORDER,
)

# 简<->繁 转换表
PAIRS = [
    ("铅", "鉛"), ("镉", "鎘"), ("亚", "亞"), ("盐", "鹽"), ("农", "農"), ("药", "藥"),
    ("残", "殘"), ("黄", "黃"), ("霉", "黴"), ("标", "標"), ("准", "準"), ("规", "規"),
    ("剂", "劑"), ("检", "檢"), ("验", "驗"), ("签", "籤"), ("过", "過"), ("敏", "敏"),
    ("总", "總"), ("数", "數"), ("肠", "腸"), ("杆", "桿"), ("门", "門"), ("干", "乾"),
    ("类", "類"), ("测", "測"), ("氯", "氯"), ("氰", "氰"), ("菊", "菊"), ("酯", "酯"),
    ("醚", "醚"), ("唑", "唑"), ("啉", "啉"), ("胺", "胺"), ("脒", "脒"), ("虫", "蟲"),
    ("噻", "噻"), ("嗪", "嗪"), ("威", "威"), ("鲜", "鮮"), ("锰", "錳"), ("吡", "吡"),
    ("嘧", "嘧"), ("灵", "靈"), ("苯", "苯"), ("环", "環"), ("啶", "啶"), ("磺", "磺"),
    ("钠", "鈉"), ("钾", "鉀"), ("钙", "鈣"), ("镁", "鎂"), ("铁", "鐵"), ("铜", "銅"), ("锌", "鋅"),
    ("镍", "鎳"), ("铬", "鉻"), ("锡", "錫"), ("汞", "汞"), ("砷", "砷"), ("硝", "硝"),
    ("芘", "芘"), ("胺", "胺"), ("联", "聯"), ("丙", "丙"), ("醇", "醇"), ("菌", "菌"),
    ("霉", "黴"), ("酵", "酵"), ("母", "母"), ("希", "希"), ("氏", "氏"),
]
SIMP2TRAD = {s: t for s, t in PAIRS}
TRAD2SIMP = {t: s for s, t in SIMP2TRAD.items()}

# ====== 全量指标清单（5 大类 72 项，按类别有序）======
# (类别, 简中名, 英文名, 关联代码, 标准框架)
ALL_INDICATORS = [
    # A. 食品添加剂（GB 2760-2024）
    ("食品添加剂", "二氧化硫/SO₂残留", "Sulfur dioxide / Sulphite residues (SO₂)", "INS E220", "GB 2760-2024"),
    ("食品添加剂", "焦亚硫酸钠", "Sodium metabisulfite", "INS 223", "GB 2760-2024"),
    ("食品添加剂", "苯甲酸及其钠盐", "Benzoic acid / Sodium benzoate", "INS 210/211", "GB 2760-2024"),
    ("食品添加剂", "山梨酸及其钾盐", "Sorbic acid / Potassium sorbate", "INS 200/202", "GB 2760-2024"),
    ("食品添加剂", "丙酸及其钠盐", "Propionic acid / Sodium propionate", "INS 280/281", "GB 2760-2024"),
    ("食品添加剂", "糖精钠", "Sodium saccharin", "INS 954", "GB 2760-2024"),
    ("食品添加剂", "甜蜜素(环己基氨基磺酸钠)", "Sodium cyclamate", "INS 952", "GB 2760-2024"),
    ("食品添加剂", "安赛蜜(乙酰磺胺酸钾)", "Acesulfame potassium", "INS 950", "GB 2760-2024"),
    ("食品添加剂", "三氯蔗糖(蔗糖素)", "Sucralose", "INS 955", "GB 2760-2024"),
    ("食品添加剂", "阿斯巴甜", "Aspartame", "INS 951", "GB 2760-2024"),
    ("食品添加剂", "胭脂红", "Ponceau 4R", "INS 124", "GB 2760-2024"),
    ("食品添加剂", "柠檬黄", "Tartrazine", "INS 102", "GB 2760-2024"),
    ("食品添加剂", "日落黄", "Sunset yellow FCF", "INS 110", "GB 2760-2024"),
    ("食品添加剂", "诱惑红", "Allura red AC", "INS 129", "GB 2760-2024"),
    ("食品添加剂", "亮蓝", "Brilliant blue FCF", "INS 133", "GB 2760-2024"),
    ("食品添加剂", "赤藓红", "Erythrosine", "INS 127", "GB 2760-2024"),
    ("食品添加剂", "柠檬酸", "Citric acid", "INS 330", "GB 2760-2024"),
    ("食品添加剂", "乳酸", "Lactic acid", "INS 270", "GB 2760-2024"),
    # B. 农药残留（GB 2763-2026）
    ("农药残留", "毒死蜱", "Chlorpyrifos", "CAS 2921-88-2", "GB 2763-2026"),
    ("农药残留", "吡虫啉", "Imidacloprid", "CAS 138261-41-3", "GB 2763-2026"),
    ("农药残留", "氯氰菊酯", "Cypermethrin", "CAS 52315-07-8", "GB 2763-2026"),
    ("农药残留", "高效氯氰菊酯", "Beta-cypermethrin", "CAS 65731-84-2", "GB 2763-2026"),
    ("农药残留", "联苯菊酯", "Bifenthrin", "CAS 82657-04-3", "GB 2763-2026"),
    ("农药残留", "咪鲜胺", "Prochloraz", "CAS 67747-09-5", "GB 2763-2026"),
    ("农药残留", "啶虫脒", "Acetamiprid", "CAS 135410-20-7", "GB 2763-2026"),
    ("农药残留", "吡唑醚菌酯", "Pyraclostrobin", "CAS 175013-18-0", "GB 2763-2026"),
    ("农药残留", "噻虫嗪", "Thiamethoxam", "CAS 153719-23-4", "GB 2763-2026"),
    ("农药残留", "克百威", "Carbofuran", "CAS 1563-66-2", "GB 2763-2026"),
    ("农药残留", "多菌灵", "Carbendazim", "CAS 10605-21-7", "GB 2763-2026"),
    ("农药残留", "苯醚甲环唑", "Difenoconazole", "CAS 119446-68-3", "GB 2763-2026"),
    ("农药残留", "甲氨基阿维菌素苯甲酸盐", "Emamectin benzoate", "CAS 155569-91-8", "GB 2763-2026"),
    ("农药残留", "氟氯氰菊酯", "Cyfluthrin", "CAS 68359-37-5", "GB 2763-2026"),
    ("农药残留", "阿维菌素", "Abamectin", "CAS 71751-41-2", "GB 2763-2026"),
    ("农药残留", "氯菊酯", "Permethrin", "CAS 52645-53-1", "GB 2763-2026"),
    ("农药残留", "溴氰菊酯", "Deltamethrin", "CAS 52918-63-5", "GB 2763-2026"),
    ("农药残留", "氰戊菊酯", "Fenvalerate", "CAS 51630-58-1", "GB 2763-2026"),
    ("农药残留", "敌敌畏", "Dichlorvos", "CAS 62-73-7", "GB 2763-2026"),
    ("农药残留", "乐果", "Dimethoate", "CAS 60-51-5", "GB 2763-2026"),
    ("农药残留", "百菌清", "Chlorothalonil", "CAS 1897-45-6", "GB 2763-2026"),
    ("农药残留", "噻螨酮", "Hexythiazox", "CAS 78587-05-0", "GB 2763-2026"),
    # C. 污染物（GB 2762-2025）
    ("污染物", "铅(Pb)", "Lead (Pb)", "CAS 7439-92-1", "GB 2762-2025"),
    ("污染物", "镉(Cd)", "Cadmium (Cd)", "CAS 7440-43-9", "GB 2762-2025"),
    ("污染物", "汞(Hg)", "Mercury (Hg)", "CAS 7439-97-6", "GB 2762-2025"),
    ("污染物", "砷(As)", "Arsenic (As)", "CAS 7440-38-2", "GB 2762-2025"),
    ("污染物", "铬(Cr)", "Chromium (Cr)", "CAS 7440-47-3", "GB 2762-2025"),
    ("污染物", "锡(Sn)", "Tin (Sn)", "CAS 7440-31-5", "GB 2762-2025"),
    ("污染物", "镍(Ni)", "Nickel (Ni)", "CAS 7440-02-0", "GB 2762-2025"),
    ("污染物", "亚硝酸盐", "Nitrite", "NO₂⁻", "GB 2762-2025"),
    ("污染物", "硝酸盐", "Nitrate", "NO₃⁻", "GB 2762-2025"),
    ("污染物", "苯并[a]芘", "Benzo[a]pyrene", "CAS 50-32-8", "GB 2762-2025"),
    ("污染物", "N-二甲基亚硝胺", "N-Nitrosodimethylamine (NDMA)", "CAS 62-75-9", "GB 2762-2025"),
    ("污染物", "多氯联苯", "Polychlorinated biphenyls (PCBs)", "", "GB 2762-2025"),
    ("污染物", "3-氯-1,2-丙二醇(3-MCPD)", "3-Monochloropropane-1,2-diol", "CAS 96-24-2", "GB 2762-2025"),
    # C+. 污染物扩充（重金属/其他污染物/兽药/塑化剂等）
    ("污染物", "铝(Al)", "Aluminium (Al)", "CAS 7429-90-5", "GB 2762-2025"),
    ("污染物", "氟(F)", "Fluoride (F)", "CAS 16984-48-8", "GB 2762-2025"),
    ("污染物", "多环芳烃(PAH4)", "Polycyclic aromatic hydrocarbons (PAH4)", "", "GB 2762-2025"),
    ("污染物", "二噁英类(PCDD/F)", "Dioxins (PCDD/F)", "", "GB 2762-2025"),
    ("污染物", "邻苯二甲酸二(2-乙基己基)酯(DEHP)", "Di(2-ethylhexyl) phthalate (DEHP)", "CAS 117-81-7", "GB 31604-2016"),
    ("污染物", "邻苯二甲酸二丁酯(DBP)", "Dibutyl phthalate (DBP)", "CAS 84-74-2", "GB 31604-2016"),
    ("污染物", "邻苯二甲酸二异壬酯(DINP)", "Diisononyl phthalate (DINP)", "", "GB 31604-2016"),
    ("污染物", "双酚A(BPA)", "Bisphenol A (BPA)", "CAS 80-05-7", "GB 31604-2016"),
    ("污染物", "三聚氰胺", "Melamine", "CAS 108-78-1", "GB 29660-2013"),
    ("污染物", "孔雀石绿", "Malachite green", "CAS 569-64-2", "农业农村部公告"),
    ("污染物", "氯霉素", "Chloramphenicol", "CAS 56-75-7", "GB 29685-2013"),
    ("污染物", "硝基呋喃代谢物", "Nitrofuran metabolites (AOZ/AMOZ/SEM/AHD)", "", "GB 31650-2019"),
    ("污染物", "己烯雌酚", "Diethylstilbestrol (DES)", "CAS 56-53-1", "GB 31650-2019"),
    ("污染物", "莱克多巴胺", "Ractopamine", "CAS 97825-25-7", "GB 31650-2019"),
    # D. 真菌毒素（GB 2761）
    ("真菌毒素", "黄曲霉毒素B1", "Aflatoxin B1", "CAS 1162-65-8", "GB 2761"),
    ("真菌毒素", "黄曲霉毒素M1", "Aflatoxin M1", "CAS 6795-23-9", "GB 2761"),
    ("真菌毒素", "黄曲霉毒素(总量)", "Aflatoxin total (B1+B2+G1+G2)", "CAS —", "GB 2761"),
    ("真菌毒素", "脱氧雪腐镰刀菌烯醇(DON)", "Deoxynivalenol (DON)", "CAS 51481-10-8", "GB 2761"),
    ("真菌毒素", "展青霉素", "Patulin", "CAS 149-29-1", "GB 2761"),
    ("真菌毒素", "赭曲霉毒素A", "Ochratoxin A", "CAS 303-47-9", "GB 2761"),
    ("真菌毒素", "玉米赤霉烯酮", "Zearalenone", "CAS 17924-92-4", "GB 2761"),
    # D+. 真菌毒素扩充
    ("真菌毒素", "伏马菌素B1", "Fumonisin B1", "CAS 116355-83-0", "GB 2761"),
    ("真菌毒素", "伏马菌素(总量)", "Fumonisins total (B1+B2)", "", "GB 2761"),
    ("真菌毒素", "T-2毒素", "T-2 toxin", "CAS 21259-20-1", "GB 2761"),
    ("真菌毒素", "HT-2毒素", "HT-2 toxin", "CAS 26934-87-2", "GB 2761"),
    ("真菌毒素", "杂色曲霉素", "Sterigmatocystin", "CAS 10048-13-2", "GB 2761"),
    # E. 微生物/致病菌（GB 29921-2021 + GB 14884-2016）
    ("微生物/致病菌", "菌落总数", "Aerobic plate count (APC)", "", "GB 14884-2016"),
    ("微生物/致病菌", "大肠菌群", "Coliforms", "", "GB 14884-2016"),
    ("微生物/致病菌", "大肠杆菌", "Escherichia coli (E. coli)", "", "GB 29921-2021"),
    ("微生物/致病菌", "沙门氏菌", "Salmonella spp.", "", "GB 29921-2021"),
    ("微生物/致病菌", "金黄色葡萄球菌", "Staphylococcus aureus", "", "GB 29921-2021"),
    ("微生物/致病菌", "蜡样芽胞杆菌", "Bacillus cereus", "", "GB 14884-2016"),
    ("微生物/致病菌", "霉菌", "Moulds", "", "GB 14884-2016"),
    ("微生物/致病菌", "酵母", "Yeasts", "", "GB 14884-2016"),
    ("微生物/致病菌", "单核细胞增生李斯特氏菌", "Listeria monocytogenes", "", "GB 29921-2021"),
    ("微生物/致病菌", "致泻大肠埃希氏菌", "Diarrheagenic Escherichia coli", "", "GB 29921-2021"),
    ("微生物/致病菌", "副溶血性弧菌", "Vibrio parahaemolyticus", "", "GB 29921-2021"),
    ("微生物/致病菌", "克罗诺杆菌属", "Cronobacter spp.", "", "GB 29921-2021"),
    ("微生物/致病菌", "志贺氏菌", "Shigella spp.", "", "GB 29921-2021"),
    # E+. 微生物/致病菌扩充
    ("微生物/致病菌", "空肠弯曲菌", "Campylobacter jejuni", "", "GB 29921-2021"),
    ("微生物/致病菌", "产气荚膜梭菌", "Clostridium perfringens", "", "GB 29921-2021"),
    ("微生物/致病菌", "肉毒梭菌", "Clostridium botulinum", "", "GB 29921-2021"),
    ("微生物/致病菌", "创伤弧菌", "Vibrio vulnificus", "", "GB 29921-2021"),
    ("微生物/致病菌", "小肠结肠炎耶尔森菌", "Yersinia enterocolitica", "", "GB 29921-2021"),
    ("微生物/致病菌", "铜绿假单胞菌", "Pseudomonas aeruginosa", "", "GB 19298-2014"),
    ("微生物/致病菌", "肠球菌", "Enterococcus spp.", "", "GB 29921-2021"),
    # F. 理化指标（质量/成分指标，按品类标准 GB 5009 系列等）
    ("理化指标", "水分", "Moisture", "", "GB 5009.3"),
    ("理化指标", "灰分", "Ash", "", "GB 5009.4"),
    ("理化指标", "蛋白质", "Protein", "", "GB 5009.5"),
    ("理化指标", "脂肪(粗脂肪)", "Fat (crude fat)", "", "GB 5009.6"),
    ("理化指标", "总糖", "Total sugar", "", "GB 5009.8"),
    ("理化指标", "还原糖", "Reducing sugar", "", "GB 5009.7"),
    ("理化指标", "pH值", "pH", "", "GB 5009.237"),
    ("理化指标", "总酸", "Total acidity", "", "GB 12456"),
    ("理化指标", "酸价", "Acid value", "", "GB 5009.229"),
    ("理化指标", "过氧化值", "Peroxide value", "", "GB 5009.227"),
    ("理化指标", "羰基价", "Carbonyl value", "", "GB 5009.230"),
    ("理化指标", "极性组分", "Polar compounds", "", "GB 5009.202"),
    ("理化指标", "挥发性盐基氮(TVB-N)", "Total volatile basic nitrogen (TVB-N)", "", "GB 5009.228"),
    ("理化指标", "组胺", "Histamine", "", "GB 5009.208"),
    ("理化指标", "三甲胺氮", "Trimethylamine nitrogen", "", "GB 5009.228"),
    ("理化指标", "氯化钠(食盐)", "Sodium chloride (salt)", "", "GB 5009.44"),
    ("理化指标", "溶剂残留", "Solvent residue", "", "GB 5009.262"),
    # G. 感官指标（依具体产品标准感官要求评定，非地区性限量）
    ("感官指标", "色泽", "Color", "", "产品标准感官要求"),
    ("感官指标", "气味", "Odor / aroma", "", "产品标准感官要求"),
    ("感官指标", "滋味", "Taste / flavour", "", "产品标准感官要求"),
    ("感官指标", "组织状态", "Texture / consistency", "", "产品标准感官要求"),
    ("感官指标", "杂质", "Extraneous matter", "", "产品标准感官要求"),
    ("感官指标", "异物", "Foreign matter", "", "产品标准感官要求"),
    ("感官指标", "霉变(可见霉斑)", "Visible mould / spoilage", "", "产品标准感官要求"),
    ("感官指标", "虫蛀/活虫", "Insect damage / live insects", "", "产品标准感官要求"),
    ("感官指标", "完整率(破碎率)", "Wholeness / broken rate", "", "产品标准感官要求"),
    ("感官指标", "饱满度", "Plumpness", "", "产品标准感官要求"),
]
# 全量指标名 -> (类别, 英文名, 代码, 标准框架)
ALL_INDICATOR_MAP = {e[1]: e for e in ALL_INDICATORS}

PEST_EN = {
    "咪鲜胺": "Prochloraz", "啶虫脒": "Acetamiprid", "吡唑醚菌酯": "Pyraclostrobin",
    "噻虫嗪": "Thiamethoxam", "克百威": "Carbofuran", "多菌灵": "Carbendazim",
    "苯醚甲环唑": "Difenoconazole", "吡虫啉": "Imidacloprid", "毒死蜱": "Chlorpyrifos",
    "氯氰菊酯": "Cypermethrin", "高效氯氰菊酯": "Beta-cypermethrin", "联苯菊酯": "Bifenthrin",
    "甲氨基阿维菌素": "Emamectin benzoate", "氟氯氰菊酯": "Cyfluthrin",
}

STD_EN = {
    "GB 2760-2024": "National Food Safety Standard – Standards for Uses of Food Additives (GB 2760-2024)",
    "GB 2762-2025": "National Food Safety Standard – Maximum Levels of Contaminants in Foods (GB 2762-2025)",
    "GB 2762-2022": "National Food Safety Standard – Maximum Levels of Contaminants in Foods (GB 2762-2022)",
    "GB 2763-2026": "National Food Safety Standard – Maximum Residue Limits for Pesticides in Food (GB 2763-2026)",
    "GB 2763-2021": "National Food Safety Standard – Maximum Residue Limits for Pesticides in Food (GB 2763-2021)",
    "GB 2761": "National Food Safety Standard – Maximum Levels of Mycotoxins in Foods (GB 2761)",
    "GB 29921-2021": "National Food Safety Standard – Pathogen Limit in Pre-packaged Foods (GB 29921-2021)",
    "GB 14884-2016": "National Food Safety Standard for Candied Fruit (GB 14884-2016)",
    "GB 5009.34": "National Food Safety Standard – Determination of Sulfur Dioxide in Foods (GB 5009.34)",
    "GB 7718-2011": "National Food Safety Standard – General Rules for the Labelling of Prepackaged Foods (GB 7718-2011)",
    "GB 2763": "National Food Safety Standard – Maximum Residue Limits for Pesticides in Food (GB 2763)",
    "(EC) 1881/2006": "Commission Regulation (EC) No 1881/2006 setting maximum levels for certain contaminants in foodstuffs",
    "(EC) 1169/2011": "Regulation (EU) No 1169/2011 on the provision of food information to consumers",
    "(EC) 1333/2008": "Regulation (EC) No 1333/2008 on food additives",
    "(EC) 396/2005": "Regulation (EC) No 396/2005 on maximum residue levels of pesticides in or on food and feed of plant and animal origin",
    "21 CFR 130.9": "U.S. Code of Federal Regulations, 21 CFR 130.9 (sulfiting agents – declaration)",
    "C.R.C., c. 870": "Canada – Food and Drug Regulations (C.R.C., c. 870)",
    "ISO 5522:1981": "ISO 5522:1981 – Dried fruits – Determination of sulfur dioxide content",
}
STD_NATIVE = {
    "(EC) 1881/2006": {"EU": "Verordnung (EG) Nr. 1881/2006"},
    "(EC) 396/2005": {"EU": "Verordnung (EG) Nr. 396/2005"},
}
QUOTE_CN = {
    "(EC) 1881/2006": "干果及其制品（直接食用或作配料）中黄曲霉毒素 B1≤2.0 µg/kg，B1+B2+G1+G2 总和≤4.0 µg/kg。",
    "(EU) 1169/2011": "当二氧化硫及亚硫酸盐总量超过 10 mg/kg 或 10 mg/L（以总 SO₂ 计）时，必须在配料表中标示。",
    "(EC) 1333/2008": "E220–E228（二氧化硫及亚硫酸盐）授权用于干制水果和蔬菜，以总 SO₂ 计的最大允许量。",
    "(EC) 396/2005": "设定植物源食品农药最大残留限量；未列明具体值的农药适用默认 MRL 0.01 mg/kg。",
    "21 CFR 130.9": "可检出的亚硫酸盐指终产品中亚硫酸盐≥10 mg/kg（ppm）。",
    "40 CFR 180.507": "建立杀菌剂嘧菌酯在芒果中的残留 tolerance 为 4 ppm。",
    "C.R.C., c. 870 B.01.010.2": "若亚硫酸盐总含量≥10 mg/kg，必须在标签上标示亚硫酸盐。",
    "JECFA（apps.who.int Chemical 985）": "ADI：0–0.7 mg/kg 体重（以 SO₂ 计），为亚硫酸盐类 Group ADI。",
    "ISO 5522:1981": "方法：酸化并加热试样，以氮气流带出释放气体，吸收并氧化后，用标准氢氧化钠溶液滴定测定硫酸。",
    "WTO SPS": "SPS 协定指定粮农组织/世卫组织 Codex 食品法典为相关标准制定组织。",
}

KNOWN_PESTICIDES = set(PEST_EN.keys()) | set(PEST_EN.values())

# ====== 权威限量数据库 BASE_LIMITS ======
# 来源：Codex Alimentarius / EU 法规 / US CFR 等联网检索结果
# 结构: (指标简中名, 地区代码) -> (限量值, 来源URL, 来源名称, 置信度, 适用范围, 备注)
# 搜索结果回传后填入；空字典表示暂无内置权威数据，回退到 risks.json + 标准体系映射
BASE_LIMITS = {
    # ===== CODEX 基准 =====
    ('三氯蔗糖(蔗糖素)', 'CODEX'): ('≤1500 mg/kg (蜜饯, 以蔗糖素计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '蜜饯', ''),
    ('三聚氰胺', 'CODEX'): ('≤2.5 mg/kg (普通食品); ≤1.0 mg/kg (婴儿食品) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '普通/婴儿食品', ''),
    ('二氧化硫/SO₂残留', 'CODEX'): ('≤1000 mg/kg (干制水果); ≤500 mg/kg (坚果); ≤100 mg/kg (甜点) (CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '干制水果/坚果/甜点', 'Codex 官方干制水果限量，修正旧值1500'),
    ('亮蓝', 'CODEX'): ('≤150 mg/kg (以亮蓝计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('伏马菌素(总量)', 'CODEX'): ('≤4000 μg/kg (玉米); ≤2000 μg/kg (玉米粉) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '玉米/玉米粉', ''),
    ('伏马菌素B1', 'CODEX'): ('≤4000 μg/kg (玉米); ≤2000 μg/kg (玉米粉) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '玉米/玉米粉', ''),
    ('安赛蜜(乙酰磺胺酸钾)', 'CODEX'): ('≤350 mg/kg (以乙酰磺胺酸钾计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('展青霉素', 'CODEX'): ('≤50 μg/kg (苹果汁); ≤25 μg/kg (固体苹果制品) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '苹果制品', ''),
    ('山梨酸及其钾盐', 'CODEX'): ('≤500 mg/kg (干果, 以山梨酸计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '干果', ''),
    ('日落黄', 'CODEX'): ('≤50 mg/kg (以日落黄计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('柠檬黄', 'CODEX'): ('≤150 mg/kg (以柠檬黄计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('焦亚硫酸钠', 'CODEX'): ('≤1000 mg/kg (以SO₂计, 干制水果, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '干制水果', ''),
    ('甜蜜素(环己基氨基磺酸钠)', 'CODEX'): ('≤250 mg/kg (以环己基氨基磺酸计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('砷(As)', 'CODEX'): ('≤0.20 mg/kg (无机砷, 米); ≤0.35 mg/kg (无机砷, 其他谷物) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '谷物(米)', ''),
    ('糖精钠', 'CODEX'): ('≤100 mg/kg (甜点, 以糖精计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '甜点', ''),
    ('胭脂红', 'CODEX'): ('≤50 mg/kg (以胭脂红计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('脱氧雪腐镰刀菌烯醇(DON)', 'CODEX'): ('≤2000 μg/kg (谷物); ≤1000 μg/kg (面粉); ≤200 μg/kg (婴幼儿谷物) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '谷物/面粉', ''),
    ('苯甲酸及其钠盐', 'CODEX'): ('≤800 mg/kg (干果, 以苯甲酸计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '干果', ''),
    ('诱惑红', 'CODEX'): ('≤300 mg/kg (以诱惑红计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('赭曲霉毒素A', 'CODEX'): ('≤5.0 μg/kg (谷物, CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '谷物', ''),
    ('铅(Pb)', 'CODEX'): ('≤0.10 mg/kg (水果, CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '水果', ''),
    ('锡(Sn)', 'CODEX'): ('≤250 mg/kg (罐头食品, CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '罐头食品', ''),
    ('镉(Cd)', 'CODEX'): ('≤0.10 mg/kg (谷物); ≤0.20 mg/kg (小麦); ≤0.40 mg/kg (大米) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '谷物/小麦/大米', ''),
    ('阿斯巴甜', 'CODEX'): ('≤1000 mg/kg (以阿斯巴甜计, CXS 192-1995)', 'https://www.fao.org/gsfaonline/', 'Codex CXS 192-1995 GSCTFF', '高', '一般食品', ''),
    ('黄曲霉毒素(总量)', 'CODEX'): ('≤15 μg/kg (花生); ≤10 μg/kg (即食坚果); ≤10 μg/kg (无花果干) (CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '花生/坚果/无花果干', ''),
    ('黄曲霉毒素M1', 'CODEX'): ('≤0.50 μg/kg (乳及乳制品, CXS 193-1995)', 'https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf', 'Codex CXS 193-1995', '高', '乳及乳制品', ''),
    # ===== EU 基准 =====
    ('3-氯-1,2-丙二醇(3-MCPD)', 'EU'): ('≤0.02 mg/kg (酱油, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '酱油', ''),
    ('三聚氰胺', 'EU'): ('≤2.5 mg/kg ((EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '一般食品', ''),
    ('乐果', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('二噁英类(PCDD/F)', 'EU'): ('≤3.5 pg WHO-PCDD/F-TEQ/g (鱼类, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '鱼类', ''),
    ('二氧化硫/SO₂残留', 'EU'): ('≤2000 mg/kg (杏/桃/葡萄/梅/无花果等); 其他干果≤500 mg/kg (以SO₂计)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02008R1333', 'Regulation (EC) No 1333/2008 Annex II', '高', '干制水果', 'EU添加剂法规'),
    ('克百威', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('吡唑醚菌酯', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('吡虫啉', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('咪鲜胺', 'EU'): ('≤0.03 mg/kg (芒果, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', ''),
    ('啶虫脒', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('噻虫嗪', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('噻螨酮', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('多氯联苯(PCBs)', 'EU'): ('≤0.000075 mg/kg (鱼类, 以脂肪计, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '鱼类', ''),
    ('多环芳烃(PAH4)', 'EU'): ('≤10 μg/kg (油脂, PAH4, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '中', '油脂', ''),
    ('多菌灵', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('孔雀石绿', 'EU'): ('不得检出 (MRL=0, Reg (EU) 37/2010)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02010R0037', 'Regulation (EU) No 37/2010', '高', '水产品', ''),
    ('展青霉素', 'EU'): ('≤25 μg/kg (固体苹果及苹果制品)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02006R1881', 'Regulation (EC) No 1881/2006 Annex', '高', '苹果制品(含干制苹果)', ''),
    ('己烯雌酚', 'EU'): ('不得检出 (MRL=0, Reg (EU) 37/2010)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02010R0037', 'Regulation (EU) No 37/2010', '高', '动物源食品', ''),
    ('敌敌畏', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('毒死蜱', 'EU'): ('≤0.01 mg/kg (芒果, 欧盟2020-11起禁用 MRL降至0.01)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005 附录', '高', '芒果', '2020-11起禁用'),
    ('氟氯氰菊酯', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('氯氰菊酯', 'EU'): ('≤0.7 mg/kg (芒果, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', ''),
    ('氯菊酯', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('氯霉素', 'EU'): ('不得检出 (MRL=0, Reg (EU) 37/2010)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02010R0037', 'Regulation (EU) No 37/2010', '高', '动物源食品', ''),
    ('氰戊菊酯', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('汞(Hg)', 'EU'): ('≤0.30 mg/kg (鱼类, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '鱼类', ''),
    ('溴氰菊酯', 'EU'): ('≤0.05 mg/kg (芒果, EFSA进口耐受量)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005 附录', '中', '芒果', '进口耐受量'),
    ('甲氨基阿维菌素苯甲酸盐', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('百菌清', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('砷(As)', 'EU'): ('≤0.15 mg/kg (无机砷, 米, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '米', ''),
    ('硝基呋喃代谢物', 'EU'): ('不得检出 (MRL=0, Reg (EU) 37/2010)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02010R0037', 'Regulation (EU) No 37/2010', '高', '动物源食品', ''),
    ('硝酸盐', 'EU'): ('≤3500 mg/kg (菠菜, 以硝酸盐计, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '菠菜', ''),
    ('联苯菊酯', 'EU'): ('≤0.5 mg/kg (芒果, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', ''),
    ('苯并[a]芘', 'EU'): ('≤2.0 μg/kg (一般食品, (EU) 2023/915 附件4)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915 Annex IV', '中', '一般食品(干制水果可参照)', '熏制肉/油脂等有更高专限量'),
    ('苯醚甲环唑', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('莱克多巴胺', 'EU'): ('不得检出 (MRL=0, Reg (EU) 37/2010)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02010R0037', 'Regulation (EU) No 37/2010', '高', '动物源食品', ''),
    ('赭曲霉毒素A', 'EU'): ('≤8.0 μg/kg (葡萄干/无花果干); ≤2.0 μg/kg (其他干果, 如芒果干)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02006R1881', 'Regulation (EU) 2022/1370 修订 (EC) 1881/2006', '高', '干制水果', '2023年生效'),
    ('铅(Pb)', 'EU'): ('≤0.10 mg/kg (水果, (EU) 2023/915 附件3.1.17)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915 Annex III', '高', '水果(干制水果按水果计)', ''),
    ('锡(Sn)', 'EU'): ('≤200 mg/kg (罐头食品, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '罐头食品', ''),
    ('镉(Cd)', 'EU'): ('≤0.05 mg/kg (水果); ≤0.20 mg/kg (树坚果) ((EU) 2023/915 附件3.2)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915 Annex III', '高', '水果/树坚果', ''),
    ('镍(Ni)', 'EU'): ('≤3.5 mg/kg (坚果, (EU) 2023/915)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32023R0915', 'Regulation (EU) 2023/915', '高', '坚果', ''),
    ('阿维菌素', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('高效氯氰菊酯', 'EU'): ('≤0.01 mg/kg (芒果, EU默认MRL, (EC) 396/2005)', 'https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection', '(EC) 396/2005', '高', '芒果', '默认MRL'),
    ('黄曲霉毒素(总量)', 'EU'): ('≤4 μg/kg (干制水果/坚果, 总黄曲霉毒素 B1+B2+G1+G2)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02006R1881', 'Regulation (EC) No 1881/2006 Annex sect.2', '高', '干制水果/坚果', ''),
    ('黄曲霉毒素B1', 'EU'): ('≤2.0 μg/kg (直接食用干制水果)', 'https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02006R1881', 'Regulation (EC) No 1881/2006 Annex sect.2', '高', '干制水果', ''),
    # ===== US 基准 =====
    ('二氧化硫/SO₂残留', 'US'): ('≥10 mg/kg 须在配料表声明', 'https://www.ecfr.gov/current/title-21/chapter-I/subchapter-B/part-101', '21 CFR 101.100', '高', '所有食品', '声明要求非限量'),
    ('吡唑醚菌酯', 'US'): ('≤0.6 ppm (芒果, 40 CFR 180.582)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.582', '高', '芒果', ''),
    ('吡虫啉', 'US'): ('≤1.0 ppm (芒果, 40 CFR 180.472)', 'https://www.ecfr.gov/current/title-40/chapter-I/subchapter-E/part-180/section-472', 'EPA 40 CFR 180.472', '高', '芒果', 'eCFR原文核实'),
    ('啶虫脒', 'US'): ('≤0.5 ppm (芒果, 40 CFR 180.578 作物组24B)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.578', '高', '芒果', ''),
    ('噻虫嗪', 'US'): ('≤0.40 ppm (芒果, 40 CFR 180.565)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.565', '高', '芒果', ''),
    ('展青霉素', 'US'): ('≤50 μg/kg (苹果汁, FDA行动水平)', 'https://www.ecfr.gov/current/title-21/chapter-I/subchapter-B/part-101', 'FDA Compliance Policy Guide §510.150', '中', '苹果汁', '行动水平非强制限量'),
    ('氯氰菊酯', 'US'): ('≤0.70 ppm (芒果, 40 CFR 180.418)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.418', '高', '芒果', ''),
    ('百菌清', 'US'): ('≤1.0 ppm (芒果, 40 CFR 180.275)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.275', '高', '芒果', ''),
    ('苯醚甲环唑', 'US'): ('≤0.07 ppm (芒果, 40 CFR 180.475)', 'https://www.ecfr.gov/', 'EPA 40 CFR 180.475', '高', '芒果', ''),
    ('铅(Pb)', 'US'): ('≤0.5 mg/kg (FDA行动水平)', 'https://www.fda.gov/food/compliance-enforcement-food/fda-guidance-documents-compliance-policy-guides', 'FDA Compliance Policy Guide', '中', '一般食品', '行动水平非强制限量'),
    ('黄曲霉毒素(总量)', 'US'): ('≤20 ppb (总黄曲霉毒素, 无花果≤10 ppb; FDA CPG 638.100/555.400)', 'https://www.fda.gov/food/compliance-enforcement-food/fda-guidance-documents-compliance-policy-guides', 'FDA Compliance Policy Guide 638.100/555.400', '高', '干制水果/坚果', '行动水平'),
    # ===== JP 基准 =====
    ('乐果', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('克百威', 'JP'): ('≤0.3 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('吡唑醚菌酯', 'JP'): ('≤0.01 mg/kg (一律标准, 芒果)', 'https://www.mhlw.go.jp/', '日本肯定列表制度(一律标准)', '高', '芒果', ''),
    ('吡虫啉', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('咪鲜胺', 'JP'): ('≤2 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('啶虫脒', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('噻虫嗪', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('噻螨酮', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('多菌灵', 'JP'): ('≤2 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('敌敌畏', 'JP'): ('≤0.1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('毒死蜱', 'JP'): ('≤0.05 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('氟氯氰菊酯', 'JP'): ('≤0.02 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('氯氰菊酯', 'JP'): ('≤0.03 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('氯菊酯', 'JP'): ('≤5.0 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('氰戊菊酯', 'JP'): ('≤1.0 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('溴氰菊酯', 'JP'): ('≤0.5 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('甲氨基阿维菌素苯甲酸盐', 'JP'): ('≤0.01 mg/kg (一律标准, 芒果)', 'https://www.mhlw.go.jp/', '日本肯定列表制度(一律标准)', '高', '芒果', ''),
    ('联苯菊酯', 'JP'): ('≤0.5 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('苯醚甲环唑', 'JP'): ('≤1 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    ('阿维菌素', 'JP'): ('≤0.01 mg/kg (芒果, 肯定列表)', 'https://www.mhlw.go.jp/', '日本肯定列表制度', '高', '芒果', ''),
    # ===== CN 基准 =====
    ('三氯蔗糖(蔗糖素)', 'CN'): ('≤1.5 g/kg (蜜饯类、凉果类, 以蔗糖素计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('乐果', 'CN'): ('≤0.01 mg/kg (芒果, 继承上级限量, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('二氧化硫/SO₂残留', 'CN'): ('≤0.1 g/kg (水果干类, 以SO₂残留量计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.02', '高', '水果干类(芒果干)', '蜜饯凉果≤0.35 g/kg'),
    ('亮蓝', 'CN'): ('≤0.025 g/kg (蜜饯类、凉果类, 以亮蓝计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('克百威', 'CN'): ('≤0.02 mg/kg (芒果, 以克百威及3-羟基克百威之和计, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('吡唑醚菌酯', 'CN'): ('≤0.05 mg/kg (芒果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '高', '杧果(芒果)', ''),
    ('吡虫啉', 'CN'): ('≤0.2 mg/kg (芒果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '高', '杧果(芒果)', ''),
    ('咪鲜胺', 'CN'): ('≤2 mg/kg (芒果, 咪鲜胺和咪鲜胺锰盐, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '高', '杧果(芒果)', ''),
    ('啶虫脒', 'CN'): ('≤2 mg/kg (芒果, 继承热带和亚热带类水果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('多菌灵', 'CN'): ('≤2 mg/kg (芒果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '高', '杧果(芒果)', ''),
    ('安赛蜜(乙酰磺胺酸钾)', 'CN'): ('≤0.3 g/kg (蜜饯类、凉果类, 以乙酰磺胺酸钾计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('山梨酸及其钾盐', 'CN'): ('≤0.5 g/kg (蜜饯类、凉果类, 以山梨酸计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('敌敌畏', 'CN'): ('≤0.2 mg/kg (芒果, 继承热带和亚热带类水果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('日落黄', 'CN'): ('≤0.1 g/kg (蜜饯类、凉果类, 以日落黄计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('柠檬黄', 'CN'): ('≤0.1 g/kg (蜜饯类、凉果类, 以柠檬黄计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('氯氰菊酯', 'CN'): ('≤0.7 mg/kg (芒果, GB 2763-2026 芒果专属, 以现行文本复核)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '中', '杧果(芒果)', '文献值, 建议复核'),
    ('氯菊酯', 'CN'): ('≤2 mg/kg (芒果, 以氯菊酯异构体之和计, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('氰戊菊酯', 'CN'): ('≤0.2 mg/kg (芒果, 以氰戊菊酯异构体之和计, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 热带和亚热带类水果', '高', '杧果(芒果)', ''),
    ('汞(Hg)', 'CN'): ('≤0.01 mg/kg (鲜水果, 总汞, GB 2762-2022)', 'https://www.bzxz.net/bzxz/196554.html', 'GB 2762-2022 表3', '高', '鲜水果(芒果)', '总汞'),
    ('溴氰菊酯', 'CN'): ('≤0.05 mg/kg (芒果, GB 2763-2026 芒果专属, 以现行文本复核)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '中', '杧果(芒果)', '文献值, 建议复核'),
    ('焦亚硫酸钠', 'CN'): ('≤0.1 g/kg (水果干类, 以SO₂计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.02', '高', '水果干类(芒果干)', '二氧化硫及亚硫酸盐组'),
    ('甜蜜素(环己基氨基磺酸钠)', 'CN'): ('≤8.0 g/kg (蜜饯类、凉果类, 以环己基氨基磺酸计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('糖精钠', 'CN'): ('≤5.0 g/kg (水果干类/蜜饯凉果, 以糖精计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.02', '高', '水果干类/蜜饯凉果(芒果干)', ''),
    ('胭脂红', 'CN'): ('≤0.05 g/kg (蜜饯类、凉果类, 以胭脂红计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('苯甲酸及其钠盐', 'CN'): ('≤0.5 g/kg (蜜饯类、凉果类, 以苯甲酸计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('苯醚甲环唑', 'CN'): ('≤0.2 mg/kg (芒果, GB 2763-2026)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '高', '杧果(芒果)', ''),
    ('赤藓红', 'CN'): ('≤0.05 g/kg (蜜饯类、凉果类, 以赤藓红计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', ''),
    ('铅(Pb)', 'CN'): ('≤0.10 mg/kg (鲜水果, GB 2762-2022)', 'https://www.bzxz.net/bzxz/196554.html', 'GB 2762-2022 表1', '高', '鲜水果(芒果)', '芒果干按脱水率折算'),
    ('镉(Cd)', 'CN'): ('≤0.05 mg/kg (鲜水果, GB 2762-2022)', 'https://www.bzxz.net/bzxz/196554.html', 'GB 2762-2022 表2', '高', '鲜水果(芒果)', '芒果干按脱水率折算'),
    ('阿斯巴甜', 'CN'): ('≤2.0 g/kg (蜜饯类、凉果类, 以阿斯巴甜计, GB 2760-2024)', 'https://2760.foodmate.net/', 'GB 2760-2024 表A.1 04.01.02.08.01', '高', '蜜饯凉果(芒果干)', '含苯丙氨酸须标识'),
    ('高效氯氰菊酯', 'CN'): ('≤0.7 mg/kg (芒果, GB 2763-2026 芒果专属, 以现行文本复核)', 'https://2763.foodvip.net/', 'GB 2763-2026 杧果(芒果)', '中', '杧果(芒果)', '文献值, 建议复核'),
}

# 指标 -> 默认不适用说明（本品类为蜜饯/干制水果）
NOT_APPLICABLE = {
    # 锡(Sn)主要针对罐装/金属容器食品
    "锡(Sn)": "本品类非罐装/金属容器食品，该指标通常不适用",
    # 多氯联苯主要针对油脂类
    "多氯联苯": "本品类非油脂类食品，该指标通常不适用",
    # 3-MCPD 主要针对酸水解植物蛋白/酱油
    "3-氯-1,2-丙二醇(3-MCPD)": "本品类非酸水解植物蛋白制品，该指标通常不适用",
}

# ====== 指标类别 → 国际标准框架映射（用于采纳国标注适用标准）======
# 即使无精确限量值，也标注该指标在该体系适用的标准框架 + URL
CODEX_STD_BY_CATEGORY = {
    "食品添加剂": ("CXS 192-1995", "https://www.fao.org/gsfaonline/", "Codex 食品添加剂和污染物通用标准 GSCTFF"),
    "农药残留": ("Codex MRL数据库", "https://www.fao.org/fao-who-codexalimentarius/codex-texts/dbs/pesticides/mrls/en/", "Codex 农药最大残留限量数据库"),
    "污染物": ("CXS 193-1995", "https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf", "Codex 食品和饲料中污染物和毒素通用标准"),
    "真菌毒素": ("CXS 193-1995", "https://www.fao.org/fileadmin/user_upload/agns/pdf/CXS_193e.pdf", "Codex 真菌毒素限量(含于污染物通用标准)"),
    "微生物/致病菌": ("CXC 1-1969", "https://www.fao.org/fileadmin/user_upload/agns/pdf/CXC_001e.pdf", "Codex 食品卫生通用原则"),
    "理化指标": ("Codex 商品标准", "https://www.fao.org/fao-who-codexalimentarius/codex-texts/gsisi/en/", "Codex 商品标准(理化质量指标, 按品类)"),
    "感官指标": ("CXC 1-1969", "https://www.fao.org/fileadmin/user_upload/agns/pdf/CXC_001e.pdf", "Codex 食品卫生通用原则(感官与卫生要求)"),
}
EU_STD_BY_CATEGORY = {
    "食品添加剂": ("(EC) 1333/2008", "https://ec.europa.eu/food/food-feed-portal/screen/food-additives/search", "EU 食品添加剂法规(及后续修订)"),
    "农药残留": ("(EC) 396/2005", "https://ec.europa.eu/food/plant/pesticides/eu-pesticides-database/public/?event=activesubstance.selection", "EU 农药残留法规"),
    "污染物": ("(EC) 1881/2006", "https://food.ec.europa.eu/food-safety/chemical-safety/contaminants/catalogue_en", "EU 食品污染物限量法规"),
    "真菌毒素": ("(EC) 1881/2006", "https://food.ec.europa.eu/food-safety/chemical-safety/contaminants/catalogue_en", "EU 真菌毒素限量(含于污染物法规)"),
    "微生物/致病菌": ("(EC) 2073/2005", "https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02005R2073", "EU 食品微生物标准"),
    "理化指标": ("(EU) 2017/2158 等", "https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:32017R2158", "EU 食品质量/减损相关法规(按品类)"),
    "感官指标": ("(EC) 852/2004", "https://eur-lex.europa.eu/legal-content/EN/TXT/?uri=CELEX:02004R0852", "EU 食品卫生法规(感官与卫生要求)"),
}
# 指标简中名 -> 类别
INDICATOR_CATEGORY = {e[1]: e[0] for e in ALL_INDICATORS}

def get_system_label(reg):
    """返回地区适用标准体系标签（用于显示）。"""
    sys = REGION_SYSTEM.get(reg, "参照CODEX")
    labels = {
        "GB": "中国国标GB体系", "EU": "欧盟法规体系", "US": "美国联邦法规体系",
        "自有": "该地区独立法规体系", "参照EU": "参照/采纳欧盟标准",
        "参照CODEX": "参照/采纳Codex标准", "参照US": "参照美国标准", "国际": "国际组织基准",
    }
    return labels.get(sys, sys)

def resolved_cell(ind_simp, reg, source_map, risk_by_indicator):
    """
    解析矩阵单元格内容。优先级:
    1. BASE_LIMITS 内置权威数据 → 返回 (值, url, 来源, 是否数据)
    2. risks.json 用户数据 (source_map) → 返回 HYPERLINK 跳转
    3. 不适用指标 → 返回 "不适用"+说明
    4. 采纳/参照 Codex/EU/US/JP 的地区且对应基准有数据 → 直接回填具体限量值(绿色数据)
    5. 其他 → 返回 "[待填写]+该地区适用体系"
    """
    # 不适用指标
    if ind_simp in NOT_APPLICABLE:
        return ("不适用", "", NOT_APPLICABLE[ind_simp], "na")
    # 1. 内置权威数据
    bl = BASE_LIMITS.get((ind_simp, reg))
    if bl:
        return (bl[0], bl[1], bl[2], "data")
    # 2. 用户数据
    src_row = source_map.get((ind_simp, reg))
    if src_row:
        return (src_row, "", "", "link")
    # 3. 标准体系映射：根据地区适用体系标注标准框架(即使无精确值)
    sys = REGION_SYSTEM.get(reg, "参照CODEX")
    cat = INDICATOR_CATEGORY.get(ind_simp, "")

    # 感官指标：依具体产品标准感官要求评定，非地区性限量
    if cat == "感官指标":
        return ("依产品标准评定", "", "感官指标(色泽/气味/滋味/组织状态/杂质/异物/霉变/虫蛀/完整率/饱满度)依具体产品标准感官要求评定，非地区性数值限量", "na")

    # 采纳Codex的地区 → 直接回填Codex具体限量值(绿色数据)
    if sys == "参照CODEX" or reg == "CODEX":
        codex_bl = BASE_LIMITS.get((ind_simp, "CODEX"))
        if codex_bl:
            return (codex_bl[0], codex_bl[1], codex_bl[2], "data")
        codex_info = CODEX_STD_BY_CATEGORY.get(cat)
        if codex_info:
            std_no, url, std_name = codex_info
            return (f"适用Codex标准框架({std_no})", url, std_name, "codex_ref")

    # EU成员国 → 适用EU法规(有精确值用精确值，否则标注法规框架)
    if sys == "EU":
        eu_bl = BASE_LIMITS.get((ind_simp, "EU"))
        if eu_bl:
            return (eu_bl[0], eu_bl[1], eu_bl[2], "data")
        eu_info = EU_STD_BY_CATEGORY.get(cat)
        if eu_info:
            std_no, url, std_name = eu_info
            return (f"适用EU法规({std_no})", url, std_name, "eu_ref")

    # 参照EU的地区 → 直接回填EU具体限量值(绿色数据)
    if sys == "参照EU":
        eu_bl = BASE_LIMITS.get((ind_simp, "EU"))
        if eu_bl:
            return (eu_bl[0], eu_bl[1], eu_bl[2], "data")
        eu_info = EU_STD_BY_CATEGORY.get(cat)
        if eu_info:
            std_no, url, std_name = eu_info
            return (f"适用EU标准框架({std_no})", url, std_name, "eu_ref")

    # US → 有精确值用，否则标注CFR框架
    if sys == "US":
        us_bl = BASE_LIMITS.get((ind_simp, "US"))
        if us_bl:
            return (us_bl[0], us_bl[1], us_bl[2], "data")
        return ("[待填写]", "https://www.ecfr.gov/", "适用美国联邦法规CFR(21CFR/40CFR180)", "gap")

    # 参照US → 直接回填US具体限量值(绿色数据)
    if sys == "参照US":
        us_bl = BASE_LIMITS.get((ind_simp, "US"))
        if us_bl:
            return (us_bl[0], us_bl[1], us_bl[2], "data")
        return ("[待填写]", "https://www.ecfr.gov/", "适用美国联邦法规CFR(21CFR/40CFR180)", "gap")

    # 参照JP → 直接回填JP具体限量值(绿色数据)
    if sys == "参照JP":
        jp_bl = BASE_LIMITS.get((ind_simp, "JP"))
        if jp_bl:
            return (jp_bl[0], jp_bl[1], jp_bl[2], "data")
        return ("[待填写]", "", "适用日本肯定列表制度(肯定列表/一律标准)", "gap")

    # GB(中国) → 通常有risks.json数据；无数据则标注GB标准框架
    if sys == "GB":
        return ("[待填写]", "", "适用中国国标GB(GB 2760/2761/2762/2763/29921)", "gap")

    # 自有标准体系
    if sys == "自有":
        return ("适用本国食品安全法规体系(具体限量待检索补充)", "", "该地区有独立食品安全法规体系，具体限量待检索补充", "framework")

    # 国际组织
    if sys == "国际":
        return ("适用国际/区域协调基准框架(具体限量待检索补充)", "", "国际组织协调框架/基准标准", "framework")

    # 4. 其他
    sys_label = get_system_label(reg)
    return ("[待填写]", "", f"该地区适用{sys_label}，本次未检索到具体限量", "gap")



def G(v):
    return "" if v is None else str(v)

def is_cjk(s):
    return any('\u4e00' <= c <= '\u9fff' for c in (s or ""))

def to_trad(s):
    return "".join(SIMP2TRAD.get(c, c) for c in (s or ""))

def to_simp(s):
    return "".join(TRAD2SIMP.get(c, c) for c in (s or ""))

def classify_indicator(raw):
    """返回物质级简中指标名；无法识别（非具体物质）返回 None。"""
    s = to_simp(G(raw))
    low = s.lower()
    if "二氧化硫" in s or "亞硫酸" in s or "亚硫酸" in s or "sulfit" in low:
        return "二氧化硫/SO₂残留"
    if "焦亚硫酸" in s:
        return "焦亚硫酸钠"
    if "苯甲酸" in s:
        return "苯甲酸及其钠盐"
    if "山梨酸" in s:
        return "山梨酸及其钾盐"
    if "丙酸" in s and "亚硝" not in s:
        return "丙酸及其钠盐"
    if "糖精" in s:
        return "糖精钠"
    if "甜蜜素" in s or "环己基" in s:
        return "甜蜜素(环己基氨基磺酸钠)"
    if "安赛蜜" in s or "乙酰磺胺" in s:
        return "安赛蜜(乙酰磺胺酸钾)"
    if "三氯蔗糖" in s or "蔗糖素" in s:
        return "三氯蔗糖(蔗糖素)"
    if "阿斯巴甜" in s:
        return "阿斯巴甜"
    if "胭脂红" in s:
        return "胭脂红"
    if "柠檬黄" in s:
        return "柠檬黄"
    if "日落黄" in s:
        return "日落黄"
    if "诱惑红" in s:
        return "诱惑红"
    if "亮蓝" in s:
        return "亮蓝"
    if "赤藓红" in s:
        return "赤藓红"
    if "柠檬酸" in s:
        return "柠檬酸"
    if "乳酸" in s and "亚硝" not in s:
        return "乳酸"
    if "鉛" in s or "铅" in s:
        return "铅(Pb)"
    if "鎘" in s or "镉" in s:
        return "镉(Cd)"
    if "汞" in s:
        return "汞(Hg)"
    if "砷" in s:
        return "砷(As)"
    if "鉻" in s or "铬" in s:
        return "铬(Cr)"
    if "錫" in s or "锡" in s and "亚硝" not in s:
        return "锡(Sn)"
    if "鎳" in s or "镍" in s:
        return "镍(Ni)"
    if "亚硝酸盐" in s or "亞硝酸" in s:
        return "亚硝酸盐"
    if "硝酸盐" in s or "硝酸" in s and "亚" not in s:
        return "硝酸盐"
    if "苯并" in s and "芘" in s:
        return "苯并[a]芘"
    if "二甲基亚硝胺" in s or "NDMA" in s:
        return "N-二甲基亚硝胺"
    if "多氯联苯" in s or "PCB" in s:
        return "多氯联苯"
    if "3-氯" in s or "3-MCPD" in s or "氯丙二醇" in s:
        return "3-氯-1,2-丙二醇(3-MCPD)"
    if "黃曲霉" in s or "黄曲霉" in s or "aflatoxin" in low:
        if "M1" in s or "M-1" in s:
            return "黄曲霉毒素M1"
        if "总量" in s or "总黄曲" in s or "total" in low:
            return "黄曲霉毒素(总量)"
        return "黄曲霉毒素B1"
    if "脱氧雪腐" in s or "DON" in s:
        return "脱氧雪腐镰刀菌烯醇(DON)"
    if "展青霉素" in s or "patulin" in low:
        return "展青霉素"
    if "赭曲霉" in s or "ochratoxin" in low:
        return "赭曲霉毒素A"
    if "玉米赤霉" in s or "zearalenone" in low:
        return "玉米赤霉烯酮"
    if "農藥" in s or "农药" in s or "除害" in s or "pesticide" in low or "mrl" in low:
        return "农药残留(MRL)"
    if s in KNOWN_PESTICIDES:
        return s
    if "菌落" in s:
        return "菌落总数"
    if "大肠菌群" in s:
        return "大肠菌群"
    if "大肠杆菌" in s or "大肠埃希" in s:
        if "致泻" in s:
            return "致泻大肠埃希氏菌"
        return "大肠杆菌"
    if "沙门" in s:
        return "沙门氏菌"
    if "葡萄" in s and "球" in s:
        return "金黄色葡萄球菌"
    if "蜡样" in s or "芽胞" in s:
        return "蜡样芽胞杆菌"
    if "黴菌" in s or "霉菌" in s:
        return "霉菌"
    if "酵母" in s:
        return "酵母"
    if "李斯特" in s or "listeria" in low:
        return "单核细胞增生李斯特氏菌"
    if "副溶血" in s:
        return "副溶血性弧菌"
    if "克罗诺" in s or "cronobacter" in low:
        return "克罗诺杆菌属"
    if "志贺" in s or "shigella" in low:
        return "志贺氏菌"
    return None

def sheet3_indicator(r):
    si = classify_indicator(r.get("indicator") or "")
    if si:
        return si
    si = classify_indicator((r.get("risk_type") or "") + " " + (r.get("quote") or ""))
    if si:
        return si
    return G(r.get("risk_type")) or "[待填写]"

def ind_multiling(ind_simp):
    """指标名四语：简中/繁中/英/官方语言。优先从全量清单取英文名。"""
    entry = ALL_INDICATOR_MAP.get(ind_simp)
    if entry:
        en = entry[2]
    else:
        en = PEST_EN.get(ind_simp, "[待填写]")
    return {"simp": ind_simp, "trad": to_trad(ind_simp), "en": en, "native": "[待填写]"}

def std_multiling(std_no, std_name, region):
    en = STD_EN.get(G(std_no).strip(), "[待填写]")
    native = STD_NATIVE.get(G(std_no).strip(), {}).get(region, "[待填写]")
    return {"zh": G(std_name) or "[待填写]", "en": en, "native": native}

def claimed_std(r):
    rt = G(r.get("risk_type"))
    limit_like = any(k in rt for k in ["限量", "MAX", "MRL", "农残", "污染物", "食品添加剂", "残留"])
    if limit_like and (G(r.get("std_no")) or G(r.get("std_name"))):
        name = G(r.get("std_name")) or ""
        no = G(r.get("std_no")) or ""
        return f"{name}（{no}）" if no else name
    return "[待填写]"

def expand_rows(risks):
    """将农药 MRL 行展开为具体物质行（仅当数据含 '中文名(值)' 序列）。"""
    out = []
    for r in risks:
        code = G(r.get("code")); rt = G(r.get("risk_type")); std = G(r.get("std_no")); quote = G(r.get("quote"))
        is_pesticide = ("农残" in code) or ("2763" in std) or ("农残" in rt) or ("农药" in rt) or ("除害" in rt)
        if is_pesticide:
            seg = quote
            m = re.search(r'[：:]\s*([^。.\n]+)', quote)
            if m:
                seg = m.group(1)
            pairs = []
            for chunk in re.split(r"[、，,；;]", seg):
                chunk = chunk.strip()
                mm = re.search(r'([一-鿿·]+?)\s*[（(]?(?:及[^）)]*)?[）)]?\s*([\d.]+)\s*(mg/kg|ppm)?', chunk)
                if not mm:
                    continue
                name = mm.group(1).strip()
                if name not in KNOWN_PESTICIDES and not (2 <= len(name) <= 8 and "农药" not in name):
                    continue
                val = mm.group(2) + ((" " + mm.group(3)) if mm.group(3) else "")
                pairs.append((name, val.strip()))
            valid = [(n, v) for n, v in pairs if n in KNOWN_PESTICIDES or (2 <= len(n) <= 8)]
            if valid:
                for name, val in valid:
                    nr = dict(r)
                    nr["indicator"] = name
                    nr["limit"] = val
                    nr["method"] = nr.get("method") or "GB 23200.113-2018（植物源性食品农药多残留测定 LC-MS/MS）"
                    nr["note"] = (G(nr.get("note")) + f"；由农药 MRL 行展开：{name}").strip("；")
                    out.append(nr)
                continue
        r = dict(r)
        r["indicator"] = sheet3_indicator(r)
        out.append(r)
    return out

def build_workbook(out_dir, food_name, basic, translations, risks, merge_mode=False):
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(os.path.join(out_dir, "standards"), exist_ok=True)
    out_path = os.path.join(out_dir, f"{food_name}_食品安全风险识别表.xlsx")
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="BBBBBB"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="0F766E"); head_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="CCFBF1"); sub_font = Font(bold=True, color="134E4A")
    cat_fill = PatternFill("solid", fgColor="F0F9FF"); cat_font = Font(bold=True, color="1E3A5F")
    gap_fill = PatternFill("solid", fgColor="FEF3C7")
    wrap = Alignment(vertical="top", wrap_text=True)

    # ===== Sheet1 基础信息（列表 + 配料多语对照） =====
    ws = wb.active; ws.title = "基础信息"
    ws.append(["字段", "内容"])
    ws.cell(1, 1).fill = head_fill; ws.cell(1, 1).font = head_font
    ws.cell(1, 2).fill = head_fill; ws.cell(1, 2).font = head_font
    base_rows = [
        ("食品中文名", basic.get("name_cn")), ("国际/英文名", basic.get("name_en") or basic.get("aliases")),
        ("别名/同义词", basic.get("aliases")), ("类别", basic.get("category")),
        ("配料", basic.get("ingredients_cn")),
        ("食品添加剂(含代码)", basic.get("additives")), ("添加剂说明", basic.get("additive_note")),
        ("过敏原", basic.get("allergens")), ("翻译置信度提示", basic.get("trans_conf")),
        ("skill 版本", SKILL_VER), ("生成日期", basic.get("gen_date")),
    ]
    for k, v in base_rows:
        ws.append([k, G(v)])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        row[0].font = Font(bold=True); row[0].border = border; row[0].alignment = wrap
        row[1].border = border; row[1].alignment = wrap
    ws.append([])
    ws.append(["配料多语对照（下列字段均按 14 种语言列出，来源 translations.json）"])
    hr = ws.max_row
    ws.cell(hr, 1).fill = sub_fill; ws.cell(hr, 1).font = sub_font
    langs = ["简中", "繁中", "英", "泰", "高棉", "越", "阿", "法", "德", "俄", "西", "葡", "日", "韩"]
    ws.append(["字段"] + langs)
    for c in range(1, len(langs) + 2):
        ws.cell(ws.max_row, c).fill = head_fill; ws.cell(ws.max_row, c).font = head_font
    tmap = {t.get("field"): t for t in (translations or [])}
    ing_fields = [f for f in tmap if f.startswith("配料")]
    for field in ["名称", "类别"] + ing_fields:
        t = tmap.get(field, {})
        ws.append([field] + [G(t.get(l)) for l in langs])
    for row in ws.iter_rows(min_row=hr + 1, max_row=ws.max_row, max_col=len(langs) + 1):
        for cell in row:
            cell.border = border; cell.alignment = wrap
    ws.column_dimensions["A"].width = 26
    for i in range(2, len(langs) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["B"].width = 40

    # ===== Sheet3 食品安全风险识别表（全量指标框架，31 列） =====
    ws3 = wb.create_sheet("食品安全风险识别表")
    cols = [
        "序号", "国家/地区(组织)代码", "中文名称", "英文名称", "本国语言名称",
        "法规名称(简中全称)", "法规名称(英全称)", "法规名称(官方语言全称)", "标准编号",
        "风险类型",
        "指标名称(简中)", "指标名称(繁中)", "指标名称(英)", "指标名称(官方语言)",
        "关联代码(INS/E/CI/CAS)", "适用对象/范围", "原文引用(本国语言)", "原文中文翻译",
        "限量值", "检测方法(标准编号/名称)", "限制条件", "声称执行标准",
        "来源文件名称", "来源网址", "网页存档路径", "命中别名/检索轨迹",
        "置信度(高/中/低)", "发布日期", "有效性", "访问日期", "备注",
    ]
    LIMIT_COL = cols.index("限量值") + 1
    ws3.append(cols)
    for c in range(1, len(cols) + 1):
        ws3.cell(1, c).fill = head_fill; ws3.cell(1, c).font = head_font

    # 去重 + 展开
    seen = set(); deduped = []
    for r in risks:
        key = (G(r.get("region")), G(r.get("std_no")), G(r.get("risk_type")), G(r.get("indicator")))
        if key in seen:
            continue
        seen.add(key); deduped.append(r)
    risks_expanded = expand_rows(deduped)

    # 按全量指标归类：指标简中名 -> [risk行列表]
    risk_by_indicator = {}
    for r in risks_expanded:
        ind = sheet3_indicator(r)
        risk_by_indicator.setdefault(ind, []).append(r)

    # 遍历全量指标清单，生成 Sheet3 行
    source_map = {}  # (指标简中名, 地区代码) -> Sheet3 行号
    r3 = 1
    seq = 0
    current_cat = None
    for cat, ind_simp, ind_en, ind_code, ind_std in ALL_INDICATORS:
        # 类别分隔行
        if cat != current_cat:
            current_cat = cat
            r3 += 1
            ws3.cell(r3, 1, f"━━ {cat} ━━")
            for c in range(1, len(cols) + 1):
                ws3.cell(r3, c).fill = cat_fill; ws3.cell(r3, c).font = cat_font

        # 找该指标的风险数据行
        matched = risk_by_indicator.get(ind_simp, [])
        if matched:
            for r in matched:
                seq += 1
                r3 += 1
                reg = G(r.get("region"))
                zh, en, nat = REGION_INFO.get(reg, (reg, reg, reg))
                quote = G(r.get("quote"))
                qcn = G(r.get("quote_cn")) or QUOTE_CN.get(G(r.get("std_no")).strip(),
                                                           "（原文为中文）" if is_cjk(quote) else "待补充")
                im = ind_multiling(ind_simp)
                sm = std_multiling(r.get("std_no"), r.get("std_name"), reg)
                claimed = claimed_std(r)
                ws3.append([
                    seq, reg, zh, en, nat,
                    sm["zh"], sm["en"], sm["native"], G(r.get("std_no")),
                    G(r.get("risk_type")),
                    im["simp"], im["trad"], im["en"], im["native"],
                    G(r.get("code")) or ind_code, G(r.get("scope")), quote, qcn,
                    G(r.get("limit")), G(r.get("method")), G(r.get("restriction")), claimed,
                    G(r.get("file")), G(r.get("url")), G(r.get("page_path")), G(r.get("hit_alias")),
                    G(r.get("confidence")), G(r.get("published")), G(r.get("validity")),
                    G(r.get("access")), G(r.get("note")),
                ])
                if G(r.get("limit")) not in ("", "[待填写]", "待补充"):
                    source_map[(ind_simp, reg)] = r3
                else:
                    # 有数据行但限量值为空，仍链接到该行（显示空值，区别于"无数据"）
                    source_map.setdefault((ind_simp, reg), r3)
        else:
            # 无数据框架行：以 CN 为框架，限量值标 [待填写]
            seq += 1
            r3 += 1
            reg = "CN"
            zh, en, nat = REGION_INFO.get(reg, (reg, reg, reg))
            im = ind_multiling(ind_simp)
            sm = std_multiling(ind_std, "", reg)
            ws3.append([
                seq, reg, zh, en, nat,
                "[待填写]", sm["en"], "[待填写]", ind_std,
                cat,
                im["simp"], im["trad"], im["en"], im["native"],
                ind_code, food_name, "", "待补充",
                "[待填写]", "[待填写]", "适用于该食品类别，本次检索未命中具体限量值", "[待填写]",
                "", "", "", "",
                "低", "", "", "", "全量指标清单框架行，待检索补充",
            ])
            # 框架行也加入 source_map（标为框架，Sheet4 链接到此行显示 [待填写]）
            if (ind_simp, reg) not in source_map:
                source_map[(ind_simp, reg)] = r3

    widths = [6, 16, 12, 14, 16, 30, 30, 26, 16, 16, 16, 16, 20, 18, 18, 22, 46, 40, 16, 30, 22, 26, 18, 38, 28, 20, 10, 14, 12, 14, 24]
    for i, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row):
        for cell in row:
            cell.border = border; cell.alignment = wrap
    ws3.freeze_panes = "C2"; ws3.auto_filter.ref = ws3.dimensions

    # ===== Sheet4 指标对比查询（全量指标 × 全量地区 按洲分区矩阵） =====
    ws4 = wb.create_sheet("指标对比查询")

    # 构建按洲分区的列序列：(类型, 内容)  类型: sep=洲分隔列, reg=地区数据列
    col_seq = []
    for cont in CONTINENT_ORDER:
        codes = REGIONS_BY_CONTINENT.get(cont, [])
        if codes:
            col_seq.append(("sep", cont))
            for reg in codes:
                col_seq.append(("reg", reg))
    n_data_cols = len(col_seq)

    # 表头行
    header = ["指标名称 \\ 国家/地区(组织)"]
    for typ, payload in col_seq:
        if typ == "sep":
            header.append(f"【{payload}】")
        else:
            header.append(REGION_INFO.get(payload, (payload, payload, payload))[0])
    ws4.append(header)
    for c in range(1, n_data_cols + 2):
        ws4.cell(1, c).fill = head_fill; ws4.cell(1, c).font = head_font
    ws4.cell(1, 1).alignment = Alignment(vertical="center", wrap_text=True)

    data_fill = PatternFill("solid", fgColor="ECFDF5")      # 绿：有数据
    ref_fill = PatternFill("solid", fgColor="FEF9C3")        # 黄：适用国际/区域标准
    na_fill = PatternFill("solid", fgColor="FEE2E2")        # 红：不适用

    current_cat4 = None
    for cat, ind_simp, ind_en, ind_code, ind_std in ALL_INDICATORS:
        if cat != current_cat4:
            current_cat4 = cat
            ws4.append([f"━━ {cat} ━━"] + [""] * n_data_cols)
            for c in range(1, n_data_cols + 2):
                ws4.cell(ws4.max_row, c).fill = cat_fill; ws4.cell(ws4.max_row, c).font = cat_font

        row = [ind_simp]
        row_kinds = []
        for typ, payload in col_seq:
            if typ == "sep":
                row.append(""); row_kinds.append("sep")
                continue
            reg = payload
            val, url, src, kind = resolved_cell(ind_simp, reg, source_map, risk_by_indicator)
            if kind == "link":
                cell_ref = f"{get_column_letter(LIMIT_COL)}{val}"
                sheet_ref = f"'食品安全风险识别表'!{cell_ref}"
                row.append(f'=HYPERLINK("#{sheet_ref}", {sheet_ref})')
            elif kind == "na":
                row.append("不适用")
            elif kind in ("data", "codex_ref", "eu_ref", "framework"):
                disp = val if len(val) <= 40 else val[:37] + "..."
                disp_safe = disp.replace('"', "'")
                if url:
                    row.append(f'=HYPERLINK("{url}", "{disp_safe}")')
                else:
                    row.append(disp)
            else:
                row.append("[待填写]")
            row_kinds.append(kind)
        ws4.append(row)
        # 着色
        for ci2, kind2 in enumerate(row_kinds, 2):
            cell = ws4.cell(ws4.max_row, ci2)
            if kind2 == "sep":
                cell.fill = cat_fill
            elif kind2 in ("data", "link"):
                cell.fill = data_fill
            elif kind2 in ("codex_ref", "eu_ref", "framework"):
                cell.fill = ref_fill
            elif kind2 == "na":
                cell.fill = na_fill

    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row):
        for cell in row:
            cell.border = border; cell.alignment = wrap
    ws4.column_dimensions["A"].width = 26
    for i in range(2, n_data_cols + 2):
        ws4.column_dimensions[get_column_letter(i)].width = 20
    ws4.freeze_panes = "B2"; ws4.auto_filter.ref = ws4.dimensions

    # 数据来源与说明
    n_codex_adopt = sum(1 for r in ALL_REGIONS if REGION_SYSTEM.get(r) == "参照CODEX")
    n_eu_adopt = sum(1 for r in ALL_REGIONS if REGION_SYSTEM.get(r) in ("EU", "参照EU"))
    gap_note_row = ws4.max_row + 2
    ws4.cell(gap_note_row, 1, "数据来源与说明")
    ws4.cell(gap_note_row, 1).font = sub_font; ws4.cell(gap_note_row, 1).fill = sub_fill
    notes = [
        f"1. 本表为全量矩阵：指标 {len(ALL_INDICATORS)} 项 × 地区 {len(ALL_REGIONS)} 个，按洲分区(国际/亚洲/欧洲/北美/南美/大洋洲/非洲)。"
        f"绿色=有数据(可点击跳转源表或外链法规原文)；黄色=适用国际/区域标准(Codex/EU)；红色=不适用；[待填写]=该地区适用相应标准体系但本次未检索到具体限量值。",
        f"2. 权威数据源：Codex Alimentarius(CXS 192-1995 GSCTFF / CXS 193-1995 污染物 / MRL数据库) / EU法规(EC 1881/2006 污染物 / EC 396/2005 农药MRL / EC 1333/2008 添加剂) / US CFR(21CFR/40CFR180)。每条数据可追溯至来源网址。",
        f"3. 标准体系映射：{n_codex_adopt}个国家/地区采纳Codex标准，{n_eu_adopt}个采纳/参照欧盟标准，各自单元格标注其适用标准体系并指向对应国际基准列。",
        "4. 农药残留MRL、微生物指标的具体限量值需按目标食品品类逐一检索；本次内置部分代表性数据(Codex/EU/US)，未覆盖项标[待填写]待后续补充。",
        "5. 本表不构成法律意见，出口/上市决策须以目标国主管机构最新发布文件为准。",
    ]
    for j, n in enumerate(notes, 1):
        ws4.cell(gap_note_row + j, 1, n)
        ws4.cell(gap_note_row + j, 1).alignment = wrap

    wb.save(out_path)
    return out_path



def main(argv):
    if len(argv) < 2:
        print("用法: python build_xlsx.py <食品目录> [<食品中文名>]")
        return 2
    food_dir = argv[1]
    if not os.path.isdir(food_dir):
        print(f"错误：目录不存在 {food_dir}")
        return 2
    basic_path = os.path.join(food_dir, "basic.json")
    trans_path = os.path.join(food_dir, "translations.json")
    risks_path = os.path.join(food_dir, "risks.json")
    if not os.path.exists(basic_path):
        print(f"错误：缺少 {basic_path}")
        return 2
    basic = json.load(open(basic_path, encoding="utf-8"))
    translations = json.load(open(trans_path, encoding="utf-8")) if os.path.exists(trans_path) else []
    risks = json.load(open(risks_path, encoding="utf-8")) if os.path.exists(risks_path) else []
    food_name = argv[2] if len(argv) > 2 else G(basic.get("name_cn")) or "食品"
    out = build_workbook(food_dir, food_name, basic, translations, risks)
    print(f"SAVED: {out}")
    print(f"风险行(去重+展开后): {len(risks)}")
    print(f"全量指标: {len(ALL_INDICATORS)} 项 × 全量地区: {len(ALL_REGIONS)} 个")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
