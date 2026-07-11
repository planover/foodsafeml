# -*- coding: utf-8 -*-
"""生成 v1.10.0 产物预览图（stats card）。读取 examples 下 xlsx，输出 PNG 到同目录。"""
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import os, sys

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "examples", "芒果干", "芒果干_食品安全风险识别表.xlsx")
OUT = os.path.join(HERE, "..", "examples", "芒果干", "产物预览_v1.10.0.png")

# ---- 量取 Sheet4 统计 ----
wb = openpyxl.load_workbook(XLSX, data_only=False)
s4 = wb["指标对比查询"]
gap = data = hyperlink = na = 0
for r in range(1, s4.max_row + 1):
    for c in range(1, s4.max_column + 1):
        v = s4.cell(r, c).value
        if isinstance(v, str) and v.strip():
            if "[待填写]" in v: gap += 1
            elif v.startswith("=HYPERLINK"): hyperlink += 1
            elif v == "不适用": na += 1
            else: data += 1
total = data + hyperlink + na + gap
gap_pct = gap / total * 100 if total else 0

# ---- 绘图 ----
W, H = 900, 560
img = Image.new("RGB", (W, H), (255, 255, 255))
d = ImageDraw.Draw(img)

def font(sz):
    # 优先尝试系统中文字体，回退默认
    for p in ["C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/simhei.ttf",
              "C:/Windows/Fonts/simsun.ttc"]:
        if os.path.exists(p):
            try: return ImageFont.truetype(p, sz)
            except Exception: pass
    return ImageFont.load_default()

title = font(30); h2 = font(20); body = font(17); big = font(40)

# 标题区
d.rectangle([0, 0, W, 90], fill=(15, 118, 110))
d.text((30, 26), "FoodSafeML v1.10.0 — 产物预览", fill=(255, 255, 255), font=title)
d.text((30, 60), "食安多语风险图鉴 · 芒果干 示例（全量指标·理化/微生物/感官）", fill=(220, 245, 240), font=body)

# 工作簿结构
d.text((30, 115), "工作簿结构（3 工作表 · 7 大类 126 项指标）", fill=(15, 118, 110), font=h2)
lines = [
    "① 基础信息 — 食品字段 + 配料多语对照(14语)",
    "② 食品安全风险识别表 — 31列源表(126项指标逐物质展开)",
    "③ 指标对比查询 — 126指标 × 235地区 按洲分区全量矩阵",
]
y = 150
for ln in lines:
    d.text((45, y), "• " + ln, fill=(40, 40, 40), font=body); y += 30

# 关键指标卡片
d.text((30, y + 10), "关键指标", fill=(15, 118, 110), font=h2)
y += 45
cards = [
    ("126 × 235", "指标 × 地区矩阵"),
    (f"{gap_pct:.1f}%", "[待填写] 占比 (v1.8.0为93.8%)"),
    (f"{hyperlink:,}", "来源可溯源外链公式"),
    (f"{data + na:,}", "实际数据值 + 不适用/依标准"),
]
cx = 30
cw = 200; ch = 90
for val, lab in cards:
    d.rectangle([cx, y, cx + cw - 10, y + ch], fill=(236, 247, 245), outline=(15, 118, 110))
    d.text((cx + 12, y + 12), val, fill=(15, 118, 110), font=big)
    d.text((cx + 12, y + 58), lab, fill=(60, 60, 60), font=body)
    cx += cw

# 底部说明
d.text((30, y + ch + 30),
       "剩余空白来自自有法规体系国家/国际组织框架/示例未提供本国数据，已如实标注[待填写]，不编造。",
       fill=(90, 90, 90), font=body)
d.text((30, y + ch + 58),
       "指标已覆盖 食品添加/农残/污染物/真毒/微生物 + 理化质量 + 感官 七大维度。",
       fill=(90, 90, 90), font=body)
d.text((30, y + ch + 58),
       "来源：Codex / EU法规 / US CFR（BASE_LIMITS 内置权威数据 + 标准体系映射）。",
       fill=(90, 90, 90), font=body)

img.save(OUT)
print("SAVED:", os.path.abspath(OUT))
